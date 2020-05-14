from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import load_workbook

wb = load_workbook('avito.xlsx')
sheet = wb[wb.sheetnames[0]]

namepricedct = {}
namepricelst = []

for i in range(1, 2):
	
	htmlcode = urlopen('https://www.avito.ru/moskva/noutbuki?p=' + str(i)).read().decode('utf-8')
	bscode = BeautifulSoup(htmlcode, features="html.parser")

	bsdiscrs = bscode.find_all('div', {'class': 'description item_table-description'})

	for bsdiscr in bsdiscrs:
		
		name = bsdiscr.find('a', {'class': 'snippet-link'}).get('title')
		namestart = name.find(' в Москве')
		name = name[:namestart:]
		
		price = str(bsdiscr.find('span', {'class': 'snippet-price'}))
		pricestart = price.find('\n')
		pricefinish = price.find('₽')
		price = price[pricestart + 1:pricefinish:]
		price = price.replace(' ', '')
		if price == 'Ценанеуказана</span':
			continue
		price = int(price)

		if namepricedct.get(name) == None:
			namepricedct[name] = price

		elif namepricedct[name] != price:
			continue
		else:
			break

for key in namepricedct:
	namepricelst.append((namepricedct[key], key))

namepricelst.sort()

for inf in namepricelst:
	print(inf)



